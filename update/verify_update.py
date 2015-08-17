import  MySQLdb
from    openpyxl       import  load_workbook
from    bmi_get_update import  *
from    rlk_mysql      import  *
import  logging




class Timer:    
    """Class for taking times"""
    def __enter__(self):
        self.start = time.clock()
        return self

    def __exit__(self, *args):
        self.end = time.clock()
        self.interval = self.end - self.start



def open_bmi():
    """Load BMI Database"""
    bmi_db = '/home/srarlk/data/bmi_Mar_2014/bmi_01_03_2014.xlsm'
    wb     =  load_workbook( filename=(bmi_db), use_iterators=True, data_only=True )
    return wb

def open_conn():
    """Load RLK Database"""
    dbase    = "reportlinker_dev"
    db       =  MySQLdb.connect( host="localhost", user="root", passwd="sraSQL", db=dbase )
    db.dbase =  dbase
    return db

def get_header_line(row):
    """Get sheet header"""
    time_seq = []
    if  row[0][3] == 'SeqID':  
        for year in  row[5:]:
            time_seq.append(int(year[3]))
    return  time_seq


def compare(list1, list2):
    equal = True
    try:
        if len(list1) != len(list2): equal = False
        for i, el in enumerate(list1):
            if el != list2[i]:  equal = False
    except:
        equal = False
    return equal


def bmi_update():
    """Main function"""

    """Set up"""
    chart      = "jnk_chart1"
    data_table = "jnk_chart_data1"

    logging.basicConfig(level=logging.DEBUG)

    wb = open_bmi()

    sheet_names  = ['Insurance']


    for sn in sheet_names:

        good = 0
        bad  = 0

        count_all_titles = 0
        count_good_titles = 0


        print "Testing results for sheet '%s'" % sn
    
        sheet     = wb.get_sheet_by_name( sn )
        data_rows = False
        time_seq  = []  

        for row  in sheet.iter_rows():

            db = open_conn()

            """Get time sequence from header; data rows start at next iteration."""
            if  not time_seq: 
                time_seq  = get_header_line(row)
            else:
                time_seq  = map(str, time_seq)
                data_rows = True

            try:
                if  data_rows:

                    # Extract info from BMI row
                    bmi  = BMI_Info(row)       
                    if bmi.title:  count_all_titles += 1

                    
                    if good_title(bmi):        

                        count_good_titles += 1

                        title_insert_good = True

                        bmi.data = bmi.get_data()

                        title_in_db = sql_title_match(db,bmi.title,table=chart)

                        if title_in_db:  
                            id_chart = title_in_db[-1]
                            sql = "select dim_1_1, value from %s where id_chart = %s order by dim_1_1 asc"\
                                  % (data_table, id_chart)
                            results = sql_results(db, sql)
                            time = []
                            data = []
                            for r in results:
                                time.append( r[0] )
                                data.append( r[1] )
                            if not compare(time, time_seq):
                                title_insert_good = False
                                bad += 1
                                print "Bad time: ", id_chart, bmi.title
                                print time
                                print time_seq
                                print 
                            if not compare(data, bmi.data):
                                title_insert_good = False
                                bad += 1
                                print "Bad data: ", id_chart, bmi.title
                                print data
                                print bmi.data
                                print 

                        else:
                            title_insert_good = False
                            bad += 1
                            print "Should be, but is not in DB:  ", bmi.title

                        if title_insert_good: good +=1 
            except Exception, e: 
                logging.exception(e)            

            db.close()           
        print "Number of successful inputs for this sheet:  ", good
        print "Total number of errors found for this sheet: ", bad
        print "Total number of titles found for this sheet: ", count_all_titles
        print "Number of good titles found for this sheet:  ", count_good_titles
        print

"""Run the code"""
with Timer() as t:
    bmi_update()

print "Total time of check: ", t.interval
 
