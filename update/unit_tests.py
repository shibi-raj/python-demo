# -*- coding: utf-8 -*-
import  MySQLdb
from    rlk_mysql      import  *
from    openpyxl       import  load_workbook
import  time
#from    bmi_quality_check import get_diff_data
from    bmi_get_update import  *
import  unicodedata
from bmi_title_cleaner import *

from collections import Counter
from itertools import chain
from unidecode import unidecode

"""Class for timing performance"""
class Timer:    
    def __enter__(self):
        self.start = time.time()
        return self

    def __exit__(self, *args):
        self.end = time.time()
        self.interval = self.end - self.start

def get_header_line(row):
    """Get sheet header"""
    time_seq = []
    if  row[0][3] == 'SeqID':  
        for year in  row[5:]:
            time_seq.append(unicode(int(year[3])))
    return  time_seq

def load_bmi():
    """Load BMI Database"""
    bmi_db = '/home/srarlk/dev/repos/bmi-update/Test_BMI_DB.xlsx'
    wb     =  load_workbook( filename=(bmi_db), use_iterators=True, data_only=True )
    return wb

def load_currency():
    currency = './currency_list.dat'
    cur_update = '/home/srarlk/data/currencies/consolidated.csv'

    with open(cur_update, 'r') as cvsfile:

        cur_ctry = []
        aux_l = []

        for line in cvsfile:
            l = line.split(',')
            if l[2].strip() != 'USD' and l[2].strip() != 'EUR' and l[2] and (l[0][0] != 'Z' and l[0][1] != 'Z')\
            and l[0][0] != '"' and l[2] != 'Alphabetic Code' and l[0].lower() != 'united states':
                if l[0].lower() == "viet nam": l[0] = "vietnam"
                aux_l.append( [ unidecode(l[2]).lower().strip(),unidecode(l[0]).lower().strip()] )

        cur_ctry.append(aux_l[0])
        for al in aux_l[1:]:
            present = False
            for i, cc in enumerate(cur_ctry):
                if  al[0] == cc[0]:
                    present = i                    
            if present:  
                cur_ctry[present].append(al[1])
            else:
                cur_ctry.append(al)

    with open(currency, "w") as internal_filename:
        pickle.dump(cur_ctry, internal_filename)




"""Open the MySQL database connection."""

dbase         = "bmi_quality"
db            =  MySQLdb.connect(host = "localhost", user = "root", passwd = "sraSQL", db = dbase)
db.dbase      =  dbase

"""BMI data object just for testing purposes."""

class Test_BMI_Info:
    def __init__(self):
        self.country     = 'Viet nam'
        self.description = ''
        self.unit       = 'VND'
        self.title       = 'Vietnam Poorest 20%,  net income per capita,  VND'
        self.source      = 'BMI'
        self.comments    = 'Some comments, ha ha ha'
        self.update_rate = 'Yearly'
        self.lst         = 'premium'
        self.bm          = 'buy'
        self.price       = '20000'
        self.id_editeur  =  1322
        self.data        =  [666]#self.get_data()

        def __repr__( self ):
            return  ' Data Model: title=%s, country=%s, description=%s, unit=%s, source=%s, comments=%s' %\
                ( self.title, self.country, self.description, self.unit, self.source, self.comments )


def bold(s):
    """Function to embolden strings output to screen."""
    return '\033[1m' + s + '\033[0m'

#_____________________________

"""BeginningBeginningBeginningBeginningBeginningBeginningBeginningBeginningBeginningBeginning 
   ofofofofofofofofofofofofofofofofofofofofofoofofofofofofofofofofofofofofofofofofofofofofoof 
   thethethethethethethethethethethethethethethethethethethethethethethethethethethethethethe 
   TestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTestsTests 
"""


def test_get_last_id():
    #print bold("get_last_id():")
    id_label = "id_chart"
    table    = "jnk_chart1"
    #print "\n  Last %s in chart1" % id_label, get_last_id(db, table, id_label)
    #print "\n\n"
    return get_last_id(db, table, id_label)

def test_get_last_id_2():
    #print bold("get_last_id_2():")
    id_label = "id_chart"
    table    = "jnk_chart1"
    #print "\n  Last %s in chart1" % id_label, get_last_id_2(db, table, id_label)
    #print "\n\n"
    return get_last_id_2(db, table, id_label)

def test_checking_cursor():
    print bold("checking cursor from sql_results():")
    title = 'United States Prescription drug sales, US$bn'
    sql     = sql_split_title( title )    
    results = sql_results( db, sql   )
    print "\n  Cursor object: ", results, "\n"
    print "  Iterate over cursor:"
    for r in results:
        print "  ",r,'\n\n'

def test_sql_split_title(title):
    print sql_split_title(title, table='jnkjnk_table')

def test_insert_chart1(db, obj, table='jnk_chart1'):
    print bold("insert_chart1()")
    return insert_chart1(db, obj, table)

def time_insert_chart1(db, obj, table='jnk_chart1'):
    pass

def time_insert_chart_data1(db):   
    print bold("time_insert_chart_data1():")
    table='jnk_chart_data1'

    """Generate fake data for insertion"""
    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)

    lower = int(100000)
    #upper = 100002
    upper = lower + 12842
    with Timer() as t:
        for id_chart in range(lower,upper):
            insert_chart_data1(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")
    return "time_insert_chart_data1: " + str(t.interval)
    
def test_insert_chart_data1_v2(db):
    print bold("time_insert_chart_data1():")
    table='jnk_chart_data1'
    id_chart = 100000001

    """Generate fake data for insertion"""
    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)

    return insert_chart_data1_v2(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")

def time_insert_chart_data1_v2(db):   
    print bold("time_insert_chart_data1_v2():")
    table='jnk_chart_data1'

    """Generate fake data for insertion"""
    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)

    lower = int(100000)
    #upper = 100002
    upper = lower + 12842
    with Timer() as t:
        for id_chart in range(lower,upper):
            insert_chart_data1_v2(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")
    return "time_insert_chart_data1_v2: " + str(t.interval)

def test_delete_data(db, table="jnk_chart_data1"):

    id_chart = 100000

    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)

    insert_chart_data1(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")
    delete_data(db, id_chart, table="jnk_chart_data1")


def time_delete_data(db, table="jnk_chart_data1"):

    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)

    lower = int(100000)
    upper = lower + 12842
    with Timer() as t:
        for id_chart in range(lower,upper):
            insert_chart_data1(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")
            delete_data(db, id_chart, table="jnk_chart_data1")
    return "time_insert_chart_data1_v2: " + str(t.interval)

def test_get_data_ids(db, data_table='jnk_chart_data1'):
    id_chart = 100000
    data_values = []
    data_time  = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)
    insert_chart_data1(db, data_table, id_chart, data_values, data_time, source="'Publisher > BMI'")
    data_ids = get_data_ids( db, id_chart, table=data_table )
    return data_ids

def test_add_it_up(db):
    sheet_name = 'test_sheet' 
    kept = 20
    new  = 10
    repl = 10
    disc = 10
    ex   = 1
    time_seq = range(1990,2024)
    return add_it_up(db, sheet_name, kept, new, repl, disc, ex, time_seq, 47743, 6537036)

def test_insert_chart_data1():
    table='jnk_chart_data1'
    id_chart=1000000
    wb = load_bmi()
    sheet_names = wb.get_sheet_names()
    time_seq = []
    for sn in sheet_names:
        sheet     = wb.get_sheet_by_name( sn )
        for row  in sheet.iter_rows():
            if  not time_seq: 
                time_seq  = get_header_line(row)
            else:
                data_rows = True
            try:
                if  data_rows:
                    bmi  = BMI_Info(row)       #  extract info from BMI row
                    bmi.fill_rest()
                    data_values = bmi.data
                    data_time = time_seq
            except: pass

    insert_chart_data1(db, table, id_chart, data_values, data_time, source="'Publisher > BMI'")
    """    
    data_values = []
    data_time   = []
    for i in range(26):
        data_values.append(i)
        data_time.append(1990 + i)
    """

def test_get_invalid_currency():
    bmi = Test_BMI_Info()
    return get_invalid_currency('Vietnam')

def test_handle_unit():
    bmi = Test_BMI_Info()
    return handle_unit(bmi)

def test_handle_currency():
    bmi = Test_BMI_Info()
    bmi.title = "Iran Health spending,  irc per capita,  2010 Prices"
    bmi.country = "Iran"
    print bmi.title
    print bmi.country
    return handle_currency(bmi)

def test_good_title():
    bmi = Test_BMI_Info()
    bmi.title = "Zimbabwe Market Orientation, JPY/U  % "
    print bmi.title
    return good_title(bmi)

def load_bmi():
    """Load BMI Database"""
    folder = '/home/srarlk/dev/repos/bmi-update/'
    bmi_file = 'Test_BMI_DB.xlsx'
    bmi_db = folder + bmi_file
    wb = load_workbook( filename=(bmi_db), use_iterators=True, data_only=True ) 
    print wb
    return wb

def get_header_line(row):
    """Get sheet header"""
    time_seq = []
    if  row[0][3] == 'SeqID':  
        for year in  row[5:]:
            time_seq.append(int(year[3]))
    return  time_seq

def test_trim_data_zeros():

    wb = load_bmi()
    sheet    = wb.get_sheet_by_name( 'Jnk_Autos' )
    time_seq = []
    data_rows = False

    i = 0
    for row in sheet.iter_rows():

        if  not time_seq: 
            time_seq  = get_header_line(row)
        else:
            data_rows = True

        if  data_rows and i < 10:
            i += 1
            bmi  = BMI_Info(row, time_seq)
            bmi.fill_rest()
            print bmi.data
            print bmi.time
            print

def test_get_diff_data(db):
    return get_diff_data(db, 2)

def test_sheets_excl():
    return sheets_excl()
    
def test_pickle_curry():
    return pickle_curry()

#_________________________________________________________


print test_good_title()

#_________________________________________________________

db.close()



