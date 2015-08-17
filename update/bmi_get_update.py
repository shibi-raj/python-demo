# -*- coding: utf-8 -*-
""" bmi_get_update """

from openpyxl           import load_workbook
from rlk_mysql          import *
import unicodedata

class BMI_Info():


    def __init__(self, row, time_seq):
        self.row         =  row
        self.time        =  time_seq[:]
        self.country     =  self.get_country()
        self.description = ''
        self.unit        = ''
        self.title       =  self.get_title()

    def __repr__( self ):
        return  ' Data Model: title=%s, country=%s, description=%s, unit=%s, source=%s, comments=%s' %\
                ( self.title, self.country, self.description, self.unit, self.source, self.comments )

    def fill_rest(self):
        self.source      =  self.get_source()
        self.comments    =  self.get_comments()
        self.update_rate = 'Yearly'
        self.lst         = 'premium'
        self.bm          = 'buy'
        self.price       = '20000'
        self.id_editeur  = '1322'
        self.data        =  self.get_data()

    def get_country(self):
        return str(self.row[1][3])

    def get_title(self):
        try:
            part = self.row[2][3].replace("'", "`")
            part = part.split(': ')
            #for p in part: p.replace("'", "`")
            if   part  and  len(part) > 1:  
                 self.description = part[0]
                 part  =        part[1]  
            elif part: 
                 part =   part[0]
            tmp_unit  = part.split(',')
            if len(tmp_unit) > 1: self.unit = tmp_unit[-1].strip()
            full_string = self.country +' '+ part
        except:
            full_string = ''
        return  self.make_uniform(full_string)
    
    def get_source(self):  #  > data source
        try:
            source = unicodedata.normalize('NFKD',self.row[3][3]).encode('ascii', 'ignore')
            source = source.replace("'", "`")
        except Exception, e:
            source = ''#unicodedata.normalize('NFKD',self.row[3][3]).encode('ascii', 'ignore')
        if str(source) == '0':  source = 'BMI'
        return source

    def get_comments(self):
        self.comments = ''
        try:
            self.comments = unicodedata.normalize('NFKD',self.row[4][3]).encode('ascii', 'ignore').replace("'", "`")
        except:
            self.comments = ''
            
    def get_data(self):
        self.data =  []
        for  d in self.row[5:]:  self.data.append(d[3])
        self.data =  [ x if x != '-' else 0. for x in self.data ]
        self.data = self.trim_data_zeros()
        return self.data

    def trim_data_zeros(self):
        for place in (0,-1):
            while self.data and self.data[place] == 0.0:
                  self.data.pop(place)
                  self.time.pop(place)
        return self.data

    def make_uniform(self, s = ''):
        ls = [ x.strip() for x in s.split(',') ]
        return ',  '.join(ls)



if __name__ == '__main__':

    """ BMI Database """

    bmi_db_loc = '/home/srarlk/data/bmi_Mar_2014/'
    bmi_db_nme = 'bmi_01_03_2014.xlsm'
    db         =  load_workbook( filename = bmi_db_loc + bmi_db_nme, use_iterators = True, data_only = True)

    sheet_name = 'Oil & Gas'
    sheet      =  db.get_sheet_by_name( sheet_name )

    for row in sheet.iter_rows():
        try:
            bm = BMI_Info( row )
            if   keep_title( bm.title ):  print bm
        except:  pass

