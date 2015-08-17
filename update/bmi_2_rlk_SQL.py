"""bmi_2_rlk_SQL.py
   See README.md for improvements to be made to the code next time this code is used.
"""

import  MySQLdb
from    openpyxl          import load_workbook
from    bmi_get_update    import *
from    rlk_mysql         import *
from    bmi_title_cleaner import *
import  time
import  sys
import  logging

logging.basicConfig(level=logging.DEBUG)


class Timer:    
    """Class for timing performance"""
    def __enter__(self):
        self.start = time.time()
        return self

    def __exit__(self, *args):
        self.end = time.time()
        self.interval = self.end - self.start

def load_bmi():
    """Load BMI Database"""
    folder = '/home/srarlk/data/bmi/'
    #folder = '/home/srarlk/dev/repos/bmi-update/'
    bmi_files = sys.argv[1:]
    wbs = []
    for bf in bmi_files:  
        bmi_db = folder + bf
        wbs.append( [load_workbook( filename=(bmi_db), use_iterators=True, data_only=True ), bf] )
    #print wbs
    return wbs


def open_conn():
    """Load RLK Database"""
    dbase    = "reportlinker_dev"
    db       =  MySQLdb.connect( host="localhost", user="root", passwd="sraSQL", db=dbase )
    db.dbase =  dbase
    return db


def get_header_line(row):
    """Get sheet header"""
    time_seq = []
    if  row[0][3] == 'SeqID':  
        for year in  row[5:]:
            time_seq.append(int(year[3]))
    return  time_seq


def bmi_update():
    """Main BMI Update function"""
    
    """Set up"""
    chart      = "chart1"
    data_table = "chart_data1"

    wbs = load_bmi()

    for wb in wbs:

        print "Current workbook: ", wb[1]

        sheet_names = wb[0].get_sheet_names()
        #sheet_names = ['GDP']

        kill_titles = []
    
        keep=0
        i=0
        j=0
        k=0
        ex=0

        no_data = 0

        for sn in sheet_names:

            print '\n\n\n\n',sn

            sheet     = wb[0].get_sheet_by_name( sn )
            time_seq  = []  
            data_rows = False

            if not any(x == sn for x in sheets_excl()):

                for row  in sheet.iter_rows():

                    db = open_conn()
    
                    #Get time sequence from header; data rows start at next iteration.
                    if  not time_seq: 
                        time_seq  = get_header_line(row)
                    else:
                        data_rows = True

                    try:
                        if  data_rows:
                            bmi  = BMI_Info(row, time_seq)       #  extract info from BMI row
    
                            if  good_title(bmi) and not any(x == bmi.title for x in kill_titles): # accept/reject title
                                bmi.fill_rest()        #  fill rest of data only for good titles
                                bmi.sa = wb[1]

                                if bmi.data:
                                    title_in_db = sql_title_match(db,bmi.title,table=chart)
                                    id_chart    = ''

                                    if not title_in_db:  #  title not already in MySQL DB, insert it
                                        id_chart = insert_chart1(db,bmi,table=chart)
                                        insert_chart_data1(db, data_table, id_chart, bmi.data, bmi.time, bmi.source)

                                    elif title_in_db:  #  title there and original
                                        that_data = get_data(db, title_in_db[-1])
                                        data_sets_the_same = equiv_data(bmi.data, that_data[0])

                                        if (bmi.sa == title_in_db[1]) and not data_sets_the_same:
                                            kill_titles.append(bmi.title)                                    
                                            delete_data(db, title_in_db[-1], table=chart)       # delete metadata
                                            delete_data(db, title_in_db[-1], table=data_table)  # delete data

                                        elif (bmi.sa != title_in_db[1]): 
                                            sql = "update %s set %s = '%s' where %s = %s" % (chart,'sa',bmi.sa,\
                                                  'id_chart',title_in_db[-1])
                                            sql_execute(db,sql)
                                            data_set_shorter = [tt for tt in that_data[1] if tt not in map(str,bmi.time)]
                                            for dss in data_set_shorter:
                                                sql = "delete from %s where id_chart = %s and dim_1_1 = %s" % (data_table,\
                                                       title_in_db[-1],dss)
                                                sql_execute(db, sql)
                                            for d,t in zip(bmi.data,bmi.time):
                                                sql  = "insert into %s (%s,%s,%s,%s,%s) " \
                                                       % (data_table,'id_chart','dim_1_1','value','value_source',\
                                                         'value_insert')
                                                sql += "values (%s,%s,%s,'%s',%s) " % (title_in_db[-1],t,d,bmi.source,\
                                                       'Now()')
                                                sql += "on duplicate key update value = %s" % d
                                                sql_execute(db,sql)
                                else:
                                    no_data += 1

                    except Exception, e: 
                        ex   += 1
                        logging.exception(e)            


                    db.close()           

            else: 
                print "excluding sheet: ", sn

            print '\n'
            #print 'kill titles                               ',kill_titles
            print '\n'
            print 'Discarded due to exceptions:     ',ex
            print 'Number without data for this DB: ',no_data
            print '\n'


#__________________________________________________________

with Timer() as t:
    bmi_update()

print "Total time elapsed: ", t.interval

 
